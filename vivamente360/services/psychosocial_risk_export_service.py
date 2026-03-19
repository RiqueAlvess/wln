"""
Serviço de Exportação de Matriz de Risco Psicossocial

Exports disponíveis:
- Excel: Planilha completa com matriz, fatores e alertas
- PGR (Word): Documento formatado para Programa de Gerenciamento de Riscos
"""

from io import BytesIO
from typing import Dict
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class PsychosocialRiskExportService:
    """Serviço de exportação de matriz de risco psicossocial"""

    # Cores por classificação
    COLORS = {
        'TRIVIAL': 'C6EFCE',        # Verde claro
        'TOLERÁVEL': 'FFEB9C',      # Amarelo claro
        'MODERADO': 'FFC7CE',       # Laranja claro
        'SUBSTANCIAL': 'F4B084',    # Vermelho claro
        'INTOLERÁVEL': 'D9D2E9',    # Roxo claro
    }

    @classmethod
    def export_to_excel(cls, avaliacao: Dict) -> Workbook:
        """
        Exporta matriz de risco para Excel.

        Args:
            avaliacao: Dict retornado por RiskAssessmentService.avaliar_campanha_completa()

        Returns:
            Workbook do openpyxl
        """
        wb = Workbook()

        # Sheet 1: Resumo Executivo
        cls._criar_sheet_resumo(wb, avaliacao)

        # Sheet 2: Matriz 5x5
        cls._criar_sheet_matriz(wb, avaliacao)

        # Sheet 3: Fatores Críticos
        cls._criar_sheet_fatores_criticos(wb, avaliacao)

        # Sheet 4: Todos os Fatores
        cls._criar_sheet_todos_fatores(wb, avaliacao)

        # Sheet 5: Alertas Críticos (se houver)
        if avaliacao.get('alertas_criticos'):
            cls._criar_sheet_alertas(wb, avaliacao)

        # Remover sheet padrão
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        return wb

    @classmethod
    def _criar_sheet_resumo(cls, wb: Workbook, avaliacao: Dict):
        """Cria sheet de resumo executivo"""
        ws = wb.active
        ws.title = "Resumo Executivo"

        campaign = avaliacao['campaign']
        empresa = avaliacao['empresa']
        resumo = avaliacao['resumo']
        matriz = avaliacao.get('matriz_ajustada', avaliacao['matriz_base'])

        # Cabeçalho
        ws['A1'] = "RESUMO EXECUTIVO - AVALIAÇÃO DE RISCOS PSICOSSOCIAIS NR-1"
        ws['A1'].font = Font(size=14, bold=True)

        # Informações gerais
        row = 3
        ws[f'A{row}'] = "Empresa:"
        ws[f'B{row}'] = empresa.nome
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Campanha:"
        ws[f'B{row}'] = campaign.nome

        row += 1
        ws[f'A{row}'] = "Data da Avaliação:"
        ws[f'B{row}'] = timezone.now().strftime('%d/%m/%Y %H:%M')

        row += 1
        ws[f'A{row}'] = "CNAE:"
        ws[f'B{row}'] = avaliacao.get('cnae', 'N/A')

        row += 1
        ws[f'A{row}'] = "Processou Análise de IA:"
        ws[f'B{row}'] = "Sim" if avaliacao.get('processou_ia') else "Não"

        # Distribuição de riscos
        row += 3
        ws[f'A{row}'] = "DISTRIBUIÇÃO DE RISCOS"
        ws[f'A{row}'].font = Font(size=12, bold=True)

        row += 1
        cls._escrever_linha_header(ws, row, ['Classificação', 'Quantidade', '%', 'Ação Necessária'])

        total = resumo.get('total_fatores', 1)

        for classificacao, label, key in [
            ('INTOLERÁVEL', '🟣 Intolerável', 'intoleraveis'),
            ('SUBSTANCIAL', '🔴 Substancial', 'substanciais'),
            ('MODERADO', '🟠 Moderado', 'moderados'),
            ('TOLERÁVEL', '🟡 Tolerável', 'toleraveis'),
            ('TRIVIAL', '🟢 Trivial', 'triviais'),
        ]:
            row += 1
            qtd = resumo.get(key, 0)
            perc = (qtd / total * 100) if total > 0 else 0

            ws[f'A{row}'] = label
            ws[f'B{row}'] = qtd
            ws[f'C{row}'] = f"{perc:.1f}%"

            # Ação baseada na classificação
            from services.risk_calculation_service import RiskCalculationService
            for (min_nr, max_nr), dados in RiskCalculationService.CLASSIFICACAO_NR.items():
                if dados['classificacao'] == classificacao:
                    ws[f'D{row}'] = dados['acao']
                    break

            # Colorir linha
            cor = cls.COLORS.get(classificacao, 'FFFFFF')
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

        # Métricas de sentimento (se houver)
        if 'sentimento_medio' in resumo:
            row += 3
            ws[f'A{row}'] = "ANÁLISE DE SENTIMENTO (IA)"
            ws[f'A{row}'].font = Font(size=12, bold=True)

            row += 1
            ws[f'A{row}'] = "Score Médio de Sentimento:"
            ws[f'B{row}'] = f"{resumo['sentimento_medio']:.2f}"

            row += 1
            ws[f'A{row}'] = "Sentimento Predominante:"
            ws[f'B{row}'] = resumo.get('sentimento_label', 'N/A')

            if 'nivel_preocupacao_predominante' in resumo:
                row += 1
                ws[f'A{row}'] = "Nível de Preocupação:"
                ws[f'B{row}'] = resumo['nivel_preocupacao_predominante']

        # Alertas críticos
        if resumo.get('alertas_criticos', 0) > 0:
            row += 3
            ws[f'A{row}'] = f"⚠️ ATENÇÃO: {resumo['alertas_criticos']} ALERTA(S) CRÍTICO(S) IDENTIFICADO(S)"
            ws[f'A{row}'].font = Font(size=12, bold=True, color="FF0000")

        # Ajustar larguras
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 60

    @classmethod
    def _criar_sheet_matriz(cls, wb: Workbook, avaliacao: Dict):
        """Cria sheet com matriz 5x5"""
        ws = wb.create_sheet("Matriz 5x5")

        ws['A1'] = "MATRIZ DE RISCO 5×5 (Probabilidade × Severidade)"
        ws['A1'].font = Font(size=14, bold=True)

        # Criar headers
        row = 3
        ws[f'A{row}'] = "Prob \\ Sev"
        ws[f'A{row}'].font = Font(bold=True)

        for s in range(1, 6):
            col = get_column_letter(s + 1)
            ws[f'{col}{row}'] = f"S={s}"
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

        # Preencher matriz
        from services.risk_calculation_service import RiskCalculationService

        for p in range(5, 0, -1):  # Inverter para começar do alto
            row += 1
            ws[f'A{row}'] = f"P={p}"
            ws[f'A{row}'].font = Font(bold=True)

            for s in range(1, 6):
                col = get_column_letter(s + 1)
                nr = p * s

                # Encontrar classificação
                classificacao = 'TRIVIAL'
                for (min_nr, max_nr), dados in RiskCalculationService.CLASSIFICACAO_NR.items():
                    if min_nr <= nr <= max_nr:
                        classificacao = dados['classificacao']
                        break

                # Escrever NR
                ws[f'{col}{row}'] = nr
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')

                # Colorir
                cor = cls.COLORS.get(classificacao, 'FFFFFF')
                ws[f'{col}{row}'].fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

        # Legenda
        row += 3
        ws[f'A{row}'] = "LEGENDA:"
        ws[f'A{row}'].font = Font(bold=True)

        for classificacao, label in [
            ('TRIVIAL', '🟢 Trivial (1-4)'),
            ('TOLERÁVEL', '🟡 Tolerável (5-9)'),
            ('MODERADO', '🟠 Moderado (10-14)'),
            ('SUBSTANCIAL', '🔴 Substancial (15-19)'),
            ('INTOLERÁVEL', '🟣 Intolerável (20-25)'),
        ]:
            row += 1
            ws[f'A{row}'] = label
            cor = cls.COLORS.get(classificacao, 'FFFFFF')
            ws[f'A{row}'].fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

        # Ajustar larguras
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    @classmethod
    def _criar_sheet_fatores_criticos(cls, wb: Workbook, avaliacao: Dict):
        """Cria sheet com fatores críticos"""
        ws = wb.create_sheet("Fatores Críticos")

        ws['A1'] = "FATORES CRÍTICOS (Intoleráveis e Substanciais)"
        ws['A1'].font = Font(size=14, bold=True)

        matriz = avaliacao.get('matriz_ajustada', avaliacao['matriz_base'])
        fatores_criticos = matriz.get('fatores_criticos', [])

        if not fatores_criticos:
            ws['A3'] = "✅ Nenhum fator crítico identificado!"
            ws['A3'].font = Font(color="00AA00", bold=True)
            return

        # Headers
        row = 3
        headers = ['Código', 'Fator de Risco', 'Categoria', 'P', 'S', 'NR', 'Classificação', 'Prazo', 'Ação Necessária']
        cls._escrever_linha_header(ws, row, headers)

        # Dados
        for fator_data in fatores_criticos:
            row += 1
            fator = fator_data.get('fator')

            ws[f'A{row}'] = fator.codigo if fator else 'N/A'
            ws[f'B{row}'] = fator.nome if fator else 'N/A'
            ws[f'C{row}'] = fator.categoria.nome if fator and fator.categoria else 'N/A'
            ws[f'D{row}'] = fator_data.get('probabilidade', 0)
            ws[f'E{row}'] = fator_data.get('severidade', 0)
            ws[f'F{row}'] = fator_data.get('nr', 0)
            ws[f'G{row}'] = fator_data.get('classificacao', 'N/A')
            ws[f'H{row}'] = fator_data.get('prazo', 'N/A')
            ws[f'I{row}'] = fator_data.get('acao', 'N/A')

            # Colorir linha
            classificacao = fator_data.get('classificacao', 'TRIVIAL')
            cor = cls.COLORS.get(classificacao, 'FFFFFF')
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws[f'{col}{row}'].fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

        # Ajustar larguras
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 60

    @classmethod
    def _criar_sheet_todos_fatores(cls, wb: Workbook, avaliacao: Dict):
        """Cria sheet com todos os fatores avaliados"""
        ws = wb.create_sheet("Todos os Fatores")

        ws['A1'] = "TODOS OS FATORES AVALIADOS"
        ws['A1'].font = Font(size=14, bold=True)

        # Headers
        row = 3
        headers = ['Dimensão HSE-IT', 'Score', 'Código', 'Fator de Risco', 'Categoria', 'P', 'S', 'NR', 'Classificação']
        cls._escrever_linha_header(ws, row, headers)

        matriz = avaliacao.get('matriz_ajustada', avaliacao['matriz_base'])

        # Dados por dimensão
        for dimensao_data in matriz.get('dimensoes', []):
            dimensao = dimensao_data.get('dimensao')
            score = dimensao_data.get('score', 0)

            for fator_data in dimensao_data.get('fatores', []):
                row += 1
                fator = fator_data.get('fator')

                ws[f'A{row}'] = dimensao.nome if dimensao else 'N/A'
                ws[f'B{row}'] = f"{score:.2f}"
                ws[f'C{row}'] = fator.codigo if fator else 'N/A'
                ws[f'D{row}'] = fator.nome if fator else 'N/A'
                ws[f'E{row}'] = fator.categoria.nome if fator and fator.categoria else 'N/A'
                ws[f'F{row}'] = fator_data.get('probabilidade', 0)
                ws[f'G{row}'] = fator_data.get('severidade', 0)
                ws[f'H{row}'] = fator_data.get('nr', 0)
                ws[f'I{row}'] = fator_data.get('classificacao', 'N/A')

                # Colorir linha
                classificacao = fator_data.get('classificacao', 'TRIVIAL')
                cor = cls.COLORS.get(classificacao, 'FFFFFF')
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')

        # Ajustar larguras
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 15

    @classmethod
    def _criar_sheet_alertas(cls, wb: Workbook, avaliacao: Dict):
        """Cria sheet com alertas críticos da IA"""
        ws = wb.create_sheet("Alertas Críticos")

        ws['A1'] = "⚠️ ALERTAS CRÍTICOS IDENTIFICADOS PELA IA"
        ws['A1'].font = Font(size=14, bold=True, color="FF0000")

        alertas = avaliacao.get('alertas_criticos', [])

        if not alertas:
            ws['A3'] = "Nenhum alerta crítico identificado."
            return

        # Headers
        row = 3
        headers = ['Tipo', 'Gravidade', 'Setor', 'Evidência', 'Recomendação Imediata', 'Encaminhamento']
        cls._escrever_linha_header(ws, row, headers)

        # Dados
        for alerta in alertas:
            row += 1

            ws[f'A{row}'] = alerta.get('tipo', 'N/A')
            ws[f'B{row}'] = alerta.get('gravidade', 'N/A')

            setor = alerta.get('setor')
            ws[f'C{row}'] = setor.nome if setor else 'N/A'

            ws[f'D{row}'] = alerta.get('evidencia', 'N/A')
            ws[f'E{row}'] = alerta.get('recomendacao_imediata', 'N/A')
            ws[f'F{row}'] = alerta.get('encaminhamento', 'N/A')

            # Colorir linha de vermelho se crítica
            if alerta.get('gravidade') == 'Crítica':
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')

        # Ajustar larguras
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 20

    @classmethod
    def _escrever_linha_header(cls, ws, row: int, headers: list):
        """Escreve linha de cabeçalho formatada"""
        for idx, header in enumerate(headers, start=1):
            col = get_column_letter(idx)
            ws[f'{col}{row}'] = header
            ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')

    @classmethod
    def export_pgr_document(cls, campaign) -> BytesIO:
        """
        Exporta relatório PGR (PDF) com uma seção por dimensão HSE-IT.
        Cada seção contém uma tabela com: Unidade, Setor, Cargo, Nível do Risco, Pontuação.

        Args:
            campaign: objeto Campaign

        Returns:
            BytesIO com o PDF gerado
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from apps.responses.models import SurveyResponse
        from apps.invitations.models import SurveyInvitation
        from services.score_service import ScoreService
        from collections import defaultdict

        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()

        style_title = ParagraphStyle(
            'PGRTitle',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=6,
            textColor=colors.HexColor('#0d3b6e'),
            alignment=TA_CENTER,
        )
        style_subtitle = ParagraphStyle(
            'PGRSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=4,
            textColor=colors.HexColor('#555555'),
            alignment=TA_CENTER,
        )
        style_section = ParagraphStyle(
            'PGRSection',
            parent=styles['Heading2'],
            fontSize=13,
            spaceBefore=16,
            spaceAfter=8,
            textColor=colors.HexColor('#0d3b6e'),
            borderPad=4,
        )
        style_normal = ParagraphStyle(
            'PGRNormal',
            parent=styles['Normal'],
            fontSize=9,
            spaceAfter=4,
        )

        # Cores de risco
        risk_colors = {
            'Aceitável':  colors.HexColor('#c6efce'),
            'Moderado':   colors.HexColor('#ffeb9c'),
            'Importante': colors.HexColor('#ffc7ce'),
            'Crítico':    colors.HexColor('#f4cccc'),
        }

        # ── Coletar dados por (unidade, setor, cargo) para cada dimensão ──
        responses = SurveyResponse.objects.filter(campaign=campaign).select_related('unidade', 'setor')
        # Cargos por (setor_id) a partir dos convites
        setor_cargos = defaultdict(set)
        for inv in SurveyInvitation.objects.filter(campaign=campaign).select_related('setor', 'cargo'):
            setor_cargos[inv.setor_id].add(inv.cargo.nome)

        # Calcular scores por (unidade, setor) para cada dimensão
        dimensao_dados = {dim: defaultdict(list) for dim in ScoreService.DIMENSOES.keys()}

        for response in responses:
            key = (response.unidade.nome, response.setor.nome, response.setor_id)
            for dimensao in ScoreService.DIMENSOES.keys():
                score = ScoreService.calcular_score_dimensao(response.respostas, dimensao)
                dimensao_dados[dimensao][key].append(score)

        empresa = campaign.empresa

        # ── Construir o documento ──
        story = []

        # Cabeçalho
        story.append(Paragraph("PROGRAMA DE GERENCIAMENTO DE RISCOS", style_title))
        story.append(Paragraph("Relatório PGR - Riscos Psicossociais (NR-1)", style_subtitle))
        story.append(Paragraph(
            f"Empresa: <b>{empresa.nome}</b> &nbsp;|&nbsp; CNPJ: {empresa.cnpj} &nbsp;|&nbsp; "
            f"Campanha: <b>{campaign.nome}</b> &nbsp;|&nbsp; "
            f"Data: {timezone.now().strftime('%d/%m/%Y')}",
            style_subtitle
        ))
        story.append(Spacer(1, 0.5 * cm))

        # Legenda
        legenda_data = [
            ['Nível', 'Faixa', 'Ação'],
            ['Aceitável', '> 3,0', 'Manter controles'],
            ['Moderado', '2,1 – 3,0', 'Monitoramento'],
            ['Importante', '1,1 – 2,0', 'Ação prioritária'],
            ['Crítico', '≤ 1,0', 'Intervenção imediata'],
        ]
        legenda_table = Table(legenda_data, colWidths=[4 * cm, 3 * cm, 8 * cm])
        legenda_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d3b6e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#c6efce')),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#ffeb9c')),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#ffc7ce')),
            ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#f4cccc')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d3b6e')),
        ]))
        story.append(legenda_table)
        story.append(Spacer(1, 0.5 * cm))

        dim_labels = {
            'demandas': 'Demandas',
            'controle': 'Controle',
            'apoio_chefia': 'Apoio da Chefia',
            'apoio_colegas': 'Apoio dos Colegas',
            'relacionamentos': 'Relacionamentos',
            'cargo': 'Cargo / Função',
            'comunicacao_mudancas': 'Comunicação e Mudanças',
        }

        for dimensao, label in dim_labels.items():
            dados = dimensao_dados.get(dimensao, {})

            story.append(Paragraph(f"Dimensão: {label}", style_section))

            if not dados:
                story.append(Paragraph("Sem dados disponíveis para esta dimensão.", style_normal))
                story.append(Spacer(1, 0.3 * cm))
                continue

            # Cabeçalho da tabela
            table_data = [['Unidade', 'Setor', 'Cargo(s)', 'Pontuação', 'Nível do Risco']]

            # Linhas ordenadas por score (pior primeiro)
            rows = []
            for (unidade_nome, setor_nome, setor_id), scores in dados.items():
                avg_score = round(sum(scores) / len(scores), 2) if scores else 0.0
                risco = ScoreService.classificar_risco(avg_score, dimensao)
                nivel_risco = risco['classificacao']

                # Mapear classificação interna para rótulo amigável
                nivel_map = {
                    'ALTO RISCO': 'Crítico',
                    'Risco Moderado': 'Importante',
                    'Risco Médio': 'Moderado',
                    'Baixo Risco': 'Aceitável',
                }
                nivel_label = nivel_map.get(nivel_risco, nivel_risco)

                cargos = ', '.join(sorted(setor_cargos.get(setor_id, []))) or '—'
                rows.append((unidade_nome, setor_nome, cargos, avg_score, nivel_label))

            # Ordenar: Crítico primeiro
            order = {'Crítico': 0, 'Importante': 1, 'Moderado': 2, 'Aceitável': 3}
            rows.sort(key=lambda r: order.get(r[4], 9))

            for unidade_nome, setor_nome, cargos, avg_score, nivel_label in rows:
                table_data.append([
                    Paragraph(unidade_nome, style_normal),
                    Paragraph(setor_nome, style_normal),
                    Paragraph(cargos, style_normal),
                    f"{avg_score:.2f}",
                    nivel_label,
                ])

            col_widths = [4.5 * cm, 3.5 * cm, 5.5 * cm, 2.5 * cm, 3 * cm]
            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

            tbl_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d3b6e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (3, 0), (4, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d3b6e')),
                ('ROWBACKGROUND', (0, 1), (-1, -1), colors.white),
            ]

            # Colorir coluna de nível do risco por linha
            for row_idx, (_, _, _, _, nivel_label) in enumerate(rows, start=1):
                bg = risk_colors.get(nivel_label, colors.white)
                tbl_style.append(('BACKGROUND', (4, row_idx), (4, row_idx), bg))

            tbl.setStyle(TableStyle(tbl_style))
            story.append(tbl)
            story.append(Spacer(1, 0.4 * cm))

        doc.build(story)
        buffer.seek(0)
        return buffer
